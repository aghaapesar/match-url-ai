import argparse
import json
import logging
import sys
import time
from pathlib import Path
from typing import List, Dict, Tuple, Optional

import pandas as pd
import xml.etree.ElementTree as ET
import yaml
from tqdm import tqdm
from urllib.parse import unquote, urlparse

from ai_client import AIClient, AIClientError
import requests
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def setup_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(stream=sys.stdout, level=level, format="%(asctime)s %(levelname)s %(message)s")


def read_config(config_path: Path) -> Dict:
    with open(config_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f) or {}


def read_old_urls(excel_path: Path, column_name: str = "url") -> List[str]:
    df = pd.read_excel(excel_path, dtype=str)
    if column_name not in df.columns:
        column_name = df.columns[0]
    urls = [str(u).strip() for u in df[column_name].dropna().tolist()]
    return [unquote(u) for u in urls]


def parse_sitemap_urls(sitemap_path: Path) -> List[str]:
    tree = ET.parse(sitemap_path)
    root = tree.getroot()
    ns = "{http://www.sitemaps.org/schemas/sitemap/0.9}"
    urls: List[str] = []
    for url_el in root.findall(f"{ns}url"):
        loc_el = url_el.find(f"{ns}loc")
        if loc_el is not None and loc_el.text:
            urls.append(unquote(loc_el.text.strip()))
    return urls


def parse_multiple_sitemaps(sitemap_paths: List[Path]) -> List[str]:
    all_urls: List[str] = []
    for p in sitemap_paths:
        try:
            urls = parse_sitemap_urls(p)
            all_urls.extend(urls)
        except Exception as e:
            logging.error("Failed to parse sitemap %s: %s", p, e)
    seen = set()
    unique_urls: List[str] = []
    for u in all_urls:
        if u not in seen:
            seen.add(u)
            unique_urls.append(u)
    return unique_urls


def fetch_single_sitemap(url: str, out_dir: Path, allow_prompt: bool = True) -> Optional[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    name = urlparse(url).path.rsplit("/", 1)[-1] or "sitemap.xml"
    if not name.endswith(".xml"):
        name = f"{name}.xml"
    path = out_dir / name
    if path.exists() and path.stat().st_size > 0:
        logging.info("Sitemap already exists, skipping download: %s", path)
        return path

    def attempt_block(attempts: int) -> bool:
        last_error = None
        for attempt in range(1, attempts + 1):
            try:
                resp = requests.get(url, timeout=60)
                resp.raise_for_status()
                path.write_bytes(resp.content)
                logging.info("Saved sitemap to %s", path)
                return True
            except Exception as e:
                last_error = e
                backoff = min(30, 1.5 ** attempt)
                logging.warning("Attempt %d/%d failed for %s: %s (retrying in %.1fs)", attempt, attempts, url, e, backoff)
                time.sleep(backoff)
        logging.error("Failed to fetch %s after %d attempts: %s", url, attempts, last_error)
        return False

    ok = attempt_block(10)
    if ok:
        return path
    if allow_prompt:
        try:
            ans = input("Download failed. Try 10 more attempts? (y/N): ").strip().lower()
        except EOFError:
            ans = "n"
        if ans in ("y", "yes"):
            ok2 = attempt_block(10)
            if ok2:
                return path
    return None


def fetch_and_save_sitemaps(urls: List[str], out_dir: Path) -> List[Path]:
    out_dir.mkdir(parents=True, exist_ok=True)
    paths: List[Path] = []
    for idx, url in enumerate(urls, start=1):
        name = urlparse(url).path.rsplit("/", 1)[-1] or f"sitemap_{idx}.xml"
        if not name.endswith(".xml"):
            name = f"{name}.xml"
        path = out_dir / name
        if path.exists() and path.stat().st_size > 0:
            logging.info("Sitemap already exists, skipping download: %s", path)
            paths.append(path)
            continue
        last_error = None
        for attempt in range(1, 11):
            try:
                resp = requests.get(url, timeout=60)
                resp.raise_for_status()
                path.write_bytes(resp.content)
                logging.info("Saved sitemap to %s", path)
                paths.append(path)
                last_error = None
                break
            except Exception as e:
                last_error = e
                backoff = min(30, 1.5 ** attempt)
                logging.warning("Attempt %d/10 failed for %s: %s (retrying in %.1fs)", attempt, url, e, backoff)
                time.sleep(backoff)
        if last_error:
            logging.error("Failed to fetch %s after 10 attempts: %s", url, last_error)
    return paths


def interactive_collect_sitemap_urls() -> List[str]:
    print("Enter sitemap URLs one by one. Type 'finishsitemaps' on a new line to continue.")
    collected: List[str] = []
    while True:
        try:
            line = input("Sitemap URL (or 'finishsitemaps'): ").strip()
        except EOFError:
            break
        if not line:
            continue
        if line.lower() == "finishsitemaps":
            break
        collected.append(line)
    return collected


def extract_slug(url: str) -> str:
    path = urlparse(url).path
    if not path:
        return ""
    parts = [p for p in path.split("/") if p]
    if not parts:
        return ""
    return parts[-1]


def extract_segments(url: str) -> List[str]:
    """Extract URL path segments (e.g., ['blog', 'post-title'] or ['shop', 'category', 'product'])"""
    path = urlparse(url).path
    if not path:
        return []
    parts = [p for p in path.split("/") if p]
    return parts


def get_primary_segment(url: str) -> str:
    """Get the primary segment (e.g., 'blog', 'shop', 'product-category')"""
    segments = extract_segments(url)
    return segments[0] if segments else ""


def is_category_or_brand_page(url: str) -> bool:
    """Check if URL is likely a category or brand page (not individual product/post)"""
    path = urlparse(url).path.lower()
    segments = extract_segments(url)
    
    # Category/brand indicators
    category_keywords = ["category", "categories", "brand", "brands", "collection", "collections", 
                         "product-category", "product-brand", "دسته", "برند", "مجموعه"]
    
    # Check if any segment contains category keywords
    for seg in segments:
        for keyword in category_keywords:
            if keyword in seg:
                return True
    
    # Heuristic: shorter paths are often categories (e.g., /shop/electronics vs /shop/electronics/phone-123)
    if len(segments) >= 2 and len(segments) <= 3:
        # Could be category level
        return True
    
    return False


def tokenize_slug(slug: str) -> List[str]:
    clean = slug.replace("-", " ").replace("_", " ")
    tokens = [t for t in clean.split() if t]
    return tokens


def heuristic_score(old_slug: str, new_slug: str) -> float:
    a = tokenize_slug(old_slug)
    b = tokenize_slug(new_slug)
    if not a or not b:
        return 0.0
    common = len(set(a) & set(b))
    denom = max(len(set(a) | set(b)), 1)
    jaccard = common / denom
    prefix = 1.0 if old_slug[:4] == new_slug[:4] else 0.0
    return 0.7 * jaccard + 0.3 * prefix


def segment_aware_score(old_url: str, new_url: str, slug_score: float) -> float:
    """Enhanced scoring that considers URL segments and prioritizes categories"""
    old_primary = get_primary_segment(old_url)
    new_primary = get_primary_segment(new_url)
    old_segments = extract_segments(old_url)
    new_segments = extract_segments(new_url)
    
    base_score = slug_score
    
    # Boost: Same primary segment (blog->blog, shop->shop)
    if old_primary and new_primary and old_primary == new_primary:
        base_score += 0.3
    
    # Boost: Category/brand pages get higher priority
    if is_category_or_brand_page(new_url):
        base_score += 0.25
    
    # Boost: More segment overlap
    common_segments = len(set(old_segments) & set(new_segments))
    if common_segments > 1:
        base_score += 0.1 * common_segments
    
    return min(base_score, 1.0)


def top_k_candidates(old_url: str, new_urls: List[str], k: int = 20) -> List[str]:
    """Select top candidates with segment-aware scoring and fallback hierarchy"""
    old_slug = extract_slug(old_url)
    old_primary = get_primary_segment(old_url)
    
    scored: List[Tuple[float, str]] = []
    for u in new_urls:
        slug_score = heuristic_score(old_slug, extract_slug(u))
        final_score = segment_aware_score(old_url, u, slug_score)
        scored.append((final_score, u))
    
    scored.sort(key=lambda x: x[0], reverse=True)
    
    # Ensure we have fallback URLs: same segment root, categories, and main segment
    candidates = [u for _, u in scored[:k]]
    
    # Add fallback: main segment URL (e.g., /blog, /shop) if not already included
    if old_primary:
        fallback_urls = [u for u in new_urls if get_primary_segment(u) == old_primary and len(extract_segments(u)) == 1]
        for fallback in fallback_urls[:2]:
            if fallback not in candidates:
                candidates.append(fallback)
    
    return candidates[:k]


def build_prompt(old_url: str, candidates: List[str]) -> Tuple[str, str]:
    old_segments = extract_segments(old_url)
    old_primary = get_primary_segment(old_url)
    
    system_prompt = (
        "You are a URL migration assistant for SEO redirects. Match legacy URLs to their best new URL "
        "based on topic/meaning and URL structure. Slugs may be in Persian. "
        "Always respond in JSON format with the required keys."
    )
    
    # Identify category/brand pages in candidates
    category_urls = [c for c in candidates if is_category_or_brand_page(c)]
    
    user_prompt = (
        f"Old URL: {old_url}\n"
        f"Primary segment: {old_primary or 'none'}\n"
        f"Path depth: {len(old_segments)}\n\n"
        "Candidate new URLs (pre-scored by segment similarity):\n"
        + "\n".join(f"- {c}" for c in candidates)
        + "\n\n"
        "**Matching Priority Rules:**\n"
        "1. **Exact Match**: Same topic/product/post in the same primary segment (e.g., blog→blog, shop→shop)\n"
        "2. **Category/Brand Fallback**: If exact match not found, prefer category or brand pages in same segment\n"
        "3. **Segment Root Fallback**: If no category found, use the main segment root (e.g., /shop, /blog)\n"
        "4. **Cross-segment**: Only as last resort, use different segment\n\n"
        "Additional considerations:\n"
        "- **Prioritize category/brand pages** over individual products/posts when uncertain\n"
        "- Consider semantic similarity (Persian-aware)\n"
        "- Match URL depth when possible (product→product, not product→category unless no alternative)\n\n"
        f"Category/brand URLs in candidates: {len(category_urls)}\n\n"
        "Return your response as a JSON object with exactly these keys:\n"
        "- best_new_url: (string) the selected URL from candidates\n"
        "- confidence: (number) 0 to 1\n"
        "- rationale: (string) explain which matching level was used and why"
    )
    return system_prompt, user_prompt


def ai_match(client: AIClient, old_url: str, candidates: List[str]) -> Dict[str, str]:
    system_prompt, user_prompt = build_prompt(old_url, candidates)
    try:
        result = client.chat_json(system_prompt, user_prompt, max_output_tokens=400)
        best = str(result.get("best_new_url") or "").strip()
        conf = float(result.get("confidence") or 0.0)
        rationale = str(result.get("rationale") or "").strip()
        if best not in candidates and candidates:
            best = candidates[0]
        return {"best_new_url": best, "confidence": conf, "rationale": rationale}
    except (AIClientError, ValueError) as e:
        return {"best_new_url": candidates[0] if candidates else "", "confidence": 0.0, "rationale": f"fallback: {e}"}


def run_matching(old_urls: List[str], new_urls: List[str], client: AIClient, mode: str, min_confidence: float) -> pd.DataFrame:
    rows = old_urls[:20] if mode == "test" else old_urls
    records: List[Dict[str, str]] = []
    for old_url in tqdm(rows, desc="Matching"):
        candidates = top_k_candidates(old_url, new_urls, k=20)
        match = ai_match(client, old_url, candidates)
        low_conf = match["confidence"] < float(min_confidence)
        rationale = match["rationale"]
        if low_conf:
            rationale = (rationale + f" | below_min_confidence<{min_confidence}>").strip()
        
        # Add segment info for debugging
        old_primary = get_primary_segment(old_url)
        new_primary = get_primary_segment(match["best_new_url"])
        is_category = is_category_or_brand_page(match["best_new_url"])
        
        records.append({
            "old_url": old_url,
            "old_segment": old_primary,
            "best_new_url": match["best_new_url"],
            "new_segment": new_primary,
            "is_category_page": is_category,
            "confidence": match["confidence"],
            "low_confidence": low_conf,
            "rationale": rationale,
            "candidates": json.dumps(candidates, ensure_ascii=False),
        })
        time.sleep(0.01)
    df = pd.DataFrame.from_records(records)
    return annotate_duplicates(df)


def annotate_duplicates(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["source_dup_of"] = ""
    df["dest_dup_of"] = ""

    first_idx_for_old: Dict[str, int] = {}
    first_idx_for_new: Dict[str, int] = {}

    for i, row in df.iterrows():
        old = row["old_url"]
        new = row["best_new_url"]
        if old in first_idx_for_old:
            first = first_idx_for_old[old]
            df.at[i, "source_dup_of"] = str(first + 2)
        else:
            first_idx_for_old[old] = i
        if new:
            if new in first_idx_for_new:
                first_n = first_idx_for_new[new]
                df.at[i, "dest_dup_of"] = str(first_n + 2)
            else:
                first_idx_for_new[new] = i
    return df


def save_excel_with_styles(df: pd.DataFrame, out_path: Path) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False)
    wb = load_workbook(out_path)
    ws = wb.active
    header = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    col_source = header.get("source_dup_of")
    col_dest = header.get("dest_dup_of")
    col_lowc = header.get("low_confidence")

    yellow = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        source_dup_val = ws.cell(row=row_idx, column=col_source).value if col_source else None
        dest_dup_val = ws.cell(row=row_idx, column=col_dest).value if col_dest else None
        lowc_val = ws.cell(row=row_idx, column=col_lowc).value if col_lowc else None
        fill = None
        if source_dup_val:
            fill = red
        elif dest_dup_val:
            fill = yellow
        elif lowc_val in (True, "TRUE", "True", 1):
            fill = orange
        if fill:
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(out_path)


def get_existing_sitemaps(sitemaps_dir: Path) -> List[Path]:
    """Find all XML files in sitemaps directory"""
    if not sitemaps_dir.exists():
        return []
    xml_files = list(sitemaps_dir.glob("*.xml"))
    return xml_files


def display_sitemap_stats(sitemap_paths: List[Path]) -> None:
    """Display statistics for each sitemap file"""
    print("\n" + "="*60)
    print(f"Found {len(sitemap_paths)} sitemap file(s) in directory:")
    print("="*60)
    
    total_urls = 0
    for idx, path in enumerate(sitemap_paths, start=1):
        try:
            urls = parse_sitemap_urls(path)
            count = len(urls)
            total_urls += count
            print(f"{idx}. {path.name}: {count} URLs")
        except Exception as e:
            print(f"{idx}. {path.name}: Error reading ({e})")
    
    print("="*60)
    print(f"Total unique URLs across all sitemaps will be calculated...")
    print("="*60 + "\n")


def interactive_flow(config: Dict, sitemaps_dir: Path, min_confidence: float, mode: str, out_path: Path, verbose: bool) -> None:
    print("Starting interactive URL matching wizard…")
    print("1) You'll be asked for the Excel file containing old URLs.")
    print("2) Then enter sitemap URLs one by one; type 'finishsitemaps' to start matching.")
    print("3) If you type 'finishsitemaps' immediately, we'll check the sitemaps folder for existing files.")
    print("4) We will download each sitemap (up to 10 tries, with optional extra tries).\n")

    # Excel path
    excel_path: Optional[Path] = None
    while not excel_path:
        try:
            p = input("Enter path to Excel file (e.g., old_urls.xlsx): ").strip()
        except EOFError:
            print("No input. Exiting.")
            return
        if not p:
            continue
        candidate = Path(p)
        if candidate.exists():
            excel_path = candidate
        else:
            print("File not found, try again.")

    # Collect sitemap URLs and download per URL immediately
    print("\nEnter sitemap URLs now. Type 'finishsitemaps' to proceed (or check existing sitemaps folder).")
    sitemap_paths: List[Path] = []
    first_input = True
    
    while True:
        try:
            line = input("Sitemap URL (or 'finishsitemaps'): ").strip()
        except EOFError:
            break
        if not line:
            continue
        if line.lower() == "finishsitemaps":
            # Check if this is the first input and no sitemaps downloaded yet
            if first_input and not sitemap_paths:
                print("\nChecking sitemaps folder for existing files...")
                existing_sitemaps = get_existing_sitemaps(sitemaps_dir)
                if existing_sitemaps:
                    display_sitemap_stats(existing_sitemaps)
                    try:
                        use_existing = input("Use these existing sitemaps? (Y/n): ").strip().lower()
                    except EOFError:
                        use_existing = "y"
                    if use_existing in ("", "y", "yes"):
                        sitemap_paths.extend(existing_sitemaps)
                        break
                    else:
                        print("Please enter sitemap URLs manually:")
                        first_input = False
                        continue
                else:
                    print(f"No sitemaps found in '{sitemaps_dir}' folder.")
                    print("Please enter sitemap URLs:")
                    first_input = False
                    continue
            break
        
        first_input = False
        path = fetch_single_sitemap(line, sitemaps_dir, allow_prompt=True)
        if path is not None:
            sitemap_paths.append(path)
        else:
            try:
                cont = input("Download failed. Continue with next sitemap? (y/N): ").strip().lower()
            except EOFError:
                cont = "n"
            if cont not in ("y", "yes"):
                print("Exiting.")
                return

    if not sitemap_paths:
        print("No sitemaps provided. Exiting.")
        return

    # Proceed with matching
    print("\n" + "="*60)
    print("Initializing AI client...")
    print("="*60)
    
    ai_config = config.get("ai", {})
    print(f"Provider: {ai_config.get('provider')}")
    print(f"Model: {ai_config.get('model')}")
    
    if ai_config.get('provider') == 'openai_compatible':
        base_url = ai_config.get('compatible_base_url', '')
        api_key = ai_config.get('compatible_api_key', '')
        print(f"Base URL: {base_url}")
        print(f"API Key: {'*' * (len(api_key) - 10) + api_key[-10:] if api_key else 'NOT SET'}")
    
    client = AIClient(ai_config)
    
    print("\nTesting connection...")
    if not client.test_connection():
        print("\n❌ Connection test failed! Please check your config.yaml settings.")
        print("   - Verify provider is set correctly")
        print("   - Verify API key is correct")
        print("   - Verify base URL is correct\n")
        try:
            cont = input("Continue anyway? (y/N): ").strip().lower()
        except EOFError:
            cont = "n"
        if cont not in ("y", "yes"):
            print("Exiting.")
            return
    else:
        print("✅ Connection test successful!\n")
    
    if verbose:
        logging.info("Reading old URLs…")
    old_urls = read_old_urls(excel_path)
    print(f"Read {len(old_urls)} old URLs from Excel.")
    print(f"Parsing {len(sitemap_paths)} sitemap file(s)…")
    new_urls = parse_multiple_sitemaps(sitemap_paths)
    print(f"Collected {len(new_urls)} unique new URLs.")
    print(f"Starting matching in {mode} mode…\n")

    df = run_matching(old_urls, new_urls, client, mode, min_confidence)
    save_excel_with_styles(df, out_path)
    print(f"\n✅ Done! Results saved to: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Match old URLs to new URLs using AI semantics.")
    parser.add_argument("--excel", help="Path to old_urls.xlsx")
    parser.add_argument("--sitemap", action="append", help="Path to a sitemap.xml (can be used multiple times)")
    parser.add_argument("--sitemap_url", action="append", help="Remote sitemap URL (can be used multiple times)")
    parser.add_argument("--interactive_sitemaps", action="store_true", help="Prompt for sitemap URLs interactively")
    parser.add_argument("--interactive", action="store_true", help="Start full interactive wizard (Excel + sitemaps)")
    parser.add_argument("--sitemaps_dir", default="sitemaps", help="Directory to save downloaded sitemaps")
    parser.add_argument("--config", required=True, help="Path to config.yaml")
    parser.add_argument("--out", default="matched_urls.xlsx", help="Output Excel path")
    parser.add_argument("--mode", choices=["test", "full"], default="test", help="Execution mode")
    parser.add_argument("--min_confidence", type=float, default=0.5, help="Minimum acceptable confidence to flag low-confidence matches")
    parser.add_argument("--verbose", action="store_true", help="Verbose logs")
    args = parser.parse_args()

    setup_logging(args.verbose)

    config_path = Path(args.config)
    out_path = Path(args.out)
    sitemaps_dir = Path(args.sitemaps_dir)
    config = read_config(config_path)

    if args.interactive:
        interactive_flow(config, sitemaps_dir, args.min_confidence, args.mode, out_path, args.verbose)
        return

    if not args.excel:
        logging.error("--excel is required unless using --interactive")
        sys.exit(2)

    excel_path = Path(args.excel)

    logging.info("Initializing AI client…")
    ai_config = config.get("ai", {})
    logging.info("Provider: %s, Model: %s", ai_config.get("provider"), ai_config.get("model"))
    
    client = AIClient(ai_config)
    
    logging.info("Testing connection…")
    if not client.test_connection():
        logging.error("❌ Connection test failed! Check your config.yaml")
        sys.exit(3)
    logging.info("✅ Connection test successful")

    logging.info("Reading old URLs…")
    old_urls = read_old_urls(excel_path)

    sitemap_paths: List[Path] = []
    if args.sitemap:
        for s in args.sitemap:
            sitemap_paths.append(Path(s))
    remote_urls: List[str] = []
    if args.sitemap_url:
        remote_urls.extend(args.sitemap_url)
    if args.interactive_sitemaps:
        remote_urls.extend(interactive_collect_sitemap_urls())
    if remote_urls:
        logging.info("Fetching %d sitemap URLs…", len(remote_urls))
        fetched = fetch_and_save_sitemaps(remote_urls, sitemaps_dir)
        sitemap_paths.extend(fetched)

    if not sitemap_paths:
        logging.error("No sitemaps provided. Use --sitemap, --sitemap_url, --interactive_sitemaps, or --interactive.")
        sys.exit(2)

    logging.info("Parsing %d sitemap files…", len(sitemap_paths))
    new_urls = parse_multiple_sitemaps(sitemap_paths)
    logging.info("Old URLs: %d, New URLs (unique): %d", len(old_urls), len(new_urls))

    logging.info("Running matching in %s mode…", args.mode)
    df = run_matching(old_urls, new_urls, client, args.mode, args.min_confidence)

    logging.info("Writing output to %s with styles", out_path)
    save_excel_with_styles(df, out_path)
    logging.info("Done.")


if __name__ == "__main__":
    main()
